import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime

class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

def print_banner():
    banner = f"""
{Colors.CYAN}{'='*70}
{Colors.BOLD}{Colors.HEADER}  SHOPEE MASTER FILE GENERATOR - STANDARDIZED v2.0{Colors.END}
{Colors.CYAN}{'='*70}
{Colors.BOLD}  Source Files  :  {Colors.GREEN}Source Files/{Colors.END}
{Colors.BOLD}  Output File   :  {Colors.GREEN}Shopee Failed Return to Seller (JAN-MARCH).xlsx{Colors.END}
{Colors.BOLD}  Format        :  {Colors.GREEN}Standardized across all sheets{Colors.END}
{Colors.CYAN}{'='*70}{Colors.END}
"""
    print(banner)

def extract_shop_name(filename):
    name_without_ext = os.path.splitext(filename)[0]
    match = re.match(r'^([A-Za-z0-9]+)', name_without_ext)
    if match:
        return match.group(1)
    return name_without_ext

def format_sheet_name(shop_name, target_wb):
    sheet_name = shop_name[:31]
    if sheet_name in target_wb.sheetnames:
        counter = 1
        while f"{sheet_name}_{counter}" in target_wb.sheetnames:
            counter += 1
        sheet_name = f"{sheet_name}_{counter}"
    return sheet_name

def create_standard_styles():
    """Create standardized styles for the workbook"""
    styles = {}
    
    # Header style - Bold, centered, with background
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    styles['header'] = {
        'font': header_font,
        'fill': header_fill,
        'alignment': header_alignment,
        'border': header_border
    }
    
    # Data style - Regular text
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    data_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    styles['data'] = {
        'font': data_font,
        'alignment': data_alignment,
        'border': data_border
    }
    
    # Number style - Right aligned numbers
    number_font = Font(name='Calibri', size=10)
    number_alignment = Alignment(horizontal='right', vertical='center')
    number_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    styles['number'] = {
        'font': number_font,
        'alignment': number_alignment,
        'border': number_border
    }
    
    # Currency style
    currency_font = Font(name='Calibri', size=10)
    currency_alignment = Alignment(horizontal='right', vertical='center')
    currency_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    styles['currency'] = {
        'font': currency_font,
        'alignment': currency_alignment,
        'border': currency_border,
        'number_format': '#,##0.00'
    }
    
    return styles

def detect_column_type(column_values):
    """Detect the type of data in a column"""
    # Check first 20 non-empty rows
    sample = [v for v in column_values[:20] if v is not None and v != '']
    
    if not sample:
        return 'text'
    
    # Check if it's a number
    numeric_count = 0
    date_count = 0
    currency_count = 0
    
    for val in sample:
        if isinstance(val, (int, float)):
            numeric_count += 1
            # Check if it looks like currency (has 2 decimal places)
            if isinstance(val, float) and val % 1 != 0:
                currency_count += 1
        elif isinstance(val, str):
            # Try to parse as number
            try:
                clean_val = val.replace(',', '').replace('$', '').replace('₱', '').strip()
                float(clean_val)
                numeric_count += 1
            except:
                pass
    
    # If more than 70% are numeric
    if numeric_count / len(sample) > 0.7:
        if currency_count / len(sample) > 0.3:
            return 'currency'
        return 'number'
    
    return 'text'

def standardize_worksheet(source_ws, target_ws, styles):
    """
    Copy data and formulas while applying standardized formatting
    """
    # Get all data including formulas
    max_row = source_ws.max_row
    max_col = source_ws.max_column
    
    # First pass: copy all data and formulas
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            source_cell = source_ws.cell(row=row_idx, column=col_idx)
            target_cell = target_ws.cell(row=row_idx, column=col_idx)
            
            # Copy value (preserve formulas)
            target_cell.value = source_cell.value
            
            # Copy number format if it exists (for formulas)
            if source_cell.number_format and not source_cell.number_format == 'General':
                target_cell.number_format = source_cell.number_format
    
    # Detect column types and apply standard formatting
    for col_idx in range(1, max_col + 1):
        # Get column values (first 100 rows or all if less)
        max_sample = min(max_row, 100)
        col_values = [source_ws.cell(row=row_idx, column=col_idx).value for row_idx in range(1, max_sample + 1)]
        
        col_type = detect_column_type(col_values)
        col_letter = get_column_letter(col_idx)
        
        # Set column width based on type
        if col_type == 'header':
            target_ws.column_dimensions[col_letter].width = 20
        elif col_type == 'number' or col_type == 'currency':
            target_ws.column_dimensions[col_letter].width = 15
        else:
            target_ws.column_dimensions[col_letter].width = 18
        
        # Apply formatting to each cell in column
        for row_idx in range(1, max_row + 1):
            target_cell = target_ws.cell(row=row_idx, column=col_idx)
            source_cell = source_ws.cell(row=row_idx, column=col_idx)
            
            # Apply style based on row and column type
            if row_idx == 1:
                # Header row
                target_cell.font = styles['header']['font']
                target_cell.fill = styles['header']['fill']
                target_cell.alignment = styles['header']['alignment']
                target_cell.border = styles['header']['border']
            else:
                # Data rows
                if col_type == 'number':
                    target_cell.font = styles['number']['font']
                    target_cell.alignment = styles['number']['alignment']
                    target_cell.border = styles['number']['border']
                    # Apply number format if not already formatted
                    if not target_cell.number_format or target_cell.number_format == 'General':
                        target_cell.number_format = '#,##0'
                elif col_type == 'currency':
                    target_cell.font = styles['currency']['font']
                    target_cell.alignment = styles['currency']['alignment']
                    target_cell.border = styles['currency']['border']
                    if not target_cell.number_format or target_cell.number_format == 'General':
                        target_cell.number_format = '₱#,##0.00'
                else:
                    target_cell.font = styles['data']['font']
                    target_cell.alignment = styles['data']['alignment']
                    target_cell.border = styles['data']['border']
    
    # Set row heights
    target_ws.row_dimensions[1].height = 25  # Header row
    for row_idx in range(2, max_row + 1):
        target_ws.row_dimensions[row_idx].height = 18
    
    # Copy merged cells
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    
    # Copy freeze panes
    if source_ws.freeze_panes:
        target_ws.freeze_panes = target_ws[source_ws.freeze_panes.coordinate]
    
    print(f"  {Colors.GREEN}✓{Colors.END} Standardized {Colors.BOLD}{target_ws.title}{Colors.END} "
          f"({Colors.YELLOW}{max_row:,}{Colors.END} rows × "
          f"{Colors.YELLOW}{max_col:,}{Colors.END} columns)")

def append_source_files(source_dir, output_file):
    # Get all XLSX files from source directory
    source_files = [f for f in os.listdir(source_dir) if f.endswith('.xlsx')]
    
    if not source_files:
        print(f"{Colors.RED}❌ No XLSX files found in source directory.{Colors.END}")
        return
    
    source_files.sort()
    
    print(f"\n{Colors.BOLD}📁 Found {Colors.GREEN}{len(source_files)}{Colors.END} source files{Colors.END}")
    print(f"{Colors.CYAN}{'─'*70}{Colors.END}\n")
    
    # Create master workbook
    target_wb = Workbook()
    if 'Sheet' in target_wb.sheetnames:
        target_wb.remove(target_wb['Sheet'])
    
    # Create standardized styles
    styles = create_standard_styles()
    
    # Statistics
    total_rows = 0
    total_cols = 0
    
    # Process each source file
    for idx, file in enumerate(source_files, 1):
        file_path = os.path.join(source_dir, file)
        shop_name = extract_shop_name(file)
        
        try:
            print(f"{Colors.BOLD}📄 [{idx}/{len(source_files)}]{Colors.END} Processing: {Colors.YELLOW}{file}{Colors.END}")
            
            # Load source workbook
            source_wb = load_workbook(file_path, data_only=False)
            source_ws = source_wb[source_wb.sheetnames[0]]
            
            # Create standardized sheet name
            sheet_name = format_sheet_name(shop_name, target_wb)
            
            # Create sheet with standardized formatting
            target_ws = target_wb.create_sheet(title=sheet_name)
            standardize_worksheet(source_ws, target_ws, styles)
            
            # Update statistics
            total_rows += source_ws.max_row
            total_cols = max(total_cols, source_ws.max_column)
            
            source_wb.close()
            
        except Exception as e:
            print(f"  {Colors.RED}✗ ERROR: {str(e)}{Colors.END}")
    
    # Save master file
    print(f"\n{Colors.CYAN}{'─'*70}{Colors.END}")
    print(f"{Colors.BOLD}💾 Saving master file...{Colors.END}")
    
    try:
        target_wb.save(output_file)
        file_size = os.path.getsize(output_file) / (1024 * 1024)
        
        print(f"\n{Colors.GREEN}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}✅ MASTER FILE CREATED SUCCESSFULLY!{Colors.END}")
        print(f"{Colors.GREEN}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}  📁 Output File  : {Colors.CYAN}{output_file}{Colors.END}")
        print(f"{Colors.BOLD}  📊 File Size    : {Colors.YELLOW}{file_size:.2f} MB{Colors.END}")
        print(f"{Colors.BOLD}  📋 Total Sheets : {Colors.YELLOW}{len(target_wb.sheetnames)}{Colors.END}")
        print(f"{Colors.BOLD}  📈 Total Rows   : {Colors.YELLOW}{total_rows:,}{Colors.END}")
        print(f"{Colors.BOLD}  📊 Total Cols   : {Colors.YELLOW}{total_cols:,}{Colors.END}")
        print(f"{Colors.GREEN}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}✨ Formatting standardized across all sheets!{Colors.END}\n")
        
    except Exception as e:
        print(f"{Colors.RED}❌ Error saving file: {str(e)}{Colors.END}")

def main():
    print_banner()
    
    current_dir = os.getcwd()
    source_dir = os.path.join(current_dir, 'Source Files')
    output_file = os.path.join(current_dir, 'Shopee Failed Return to Seller (JAN-MARCH).xlsx')
    
    if not os.path.exists(source_dir):
        print(f"{Colors.RED}❌ Error: Source directory '{source_dir}' not found.{Colors.END}")
        return
    
    start_time = datetime.now()
    append_source_files(source_dir, output_file)
    elapsed = datetime.now() - start_time
    print(f"{Colors.CYAN}⏱️  Total time: {Colors.YELLOW}{elapsed.total_seconds():.2f}{Colors.END} seconds\n")

if __name__ == "__main__":
    main()