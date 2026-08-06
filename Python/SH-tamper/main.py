import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

def print_banner():
    banner = f"""
{Colors.CYAN}{'='*70}
{Colors.BOLD}{Colors.HEADER}  SHOPEE TRACKING NUMBER EXTRACTOR v1.0{Colors.END}
{Colors.CYAN}{'='*70}
{Colors.BOLD}  Input File    :  {Colors.GREEN}Shopee Failed Return to Seller (JAN-MARCH).xlsx{Colors.END}
{Colors.BOLD}  Output File   :  {Colors.GREEN}Shopee Tracking Numbers (JAN-MARCH).xlsx{Colors.END}
{Colors.BOLD}  Columns       :  {Colors.GREEN}Tracking # | Date Liquidated | Logistics Remarks{Colors.END}
{Colors.CYAN}{'='*70}{Colors.END}
"""
    print(banner)

def find_tracking_column(sheet):
    """
    Find the column that contains tracking numbers.
    Priority: "Return Tracking Number" > "Tracking Number*"
    """
    # Get header row (assuming first row)
    header_row = 1
    tracking_col = None
    found_column_name = None
    
    # Get all column headers from first row
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if cell_value:
            # Clean the header text
            header_text = str(cell_value).strip()
            
            # Check for "Return Tracking Number" (case insensitive)
            if re.search(r'return\s*tracking\s*number', header_text, re.IGNORECASE):
                tracking_col = col_idx
                found_column_name = 'Return Tracking Number'
                print(f"  {Colors.GREEN}✓{Colors.END} Found '{found_column_name}' at column {get_column_letter(col_idx)}")
                return tracking_col, found_column_name
    
    # If not found, check for "Tracking Number*" (case insensitive)
    for col_idx in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if cell_value:
            header_text = str(cell_value).strip()
            if re.search(r'tracking\s*number\*?', header_text, re.IGNORECASE):
                tracking_col = col_idx
                found_column_name = 'Tracking Number*'
                print(f"  {Colors.YELLOW}⚠{Colors.END} 'Return Tracking Number' not found. Using '{found_column_name}' at column {get_column_letter(col_idx)}")
                return tracking_col, found_column_name
    
    return None, None

def extract_tracking_numbers(input_file, output_file):
    """
    Extract tracking numbers from each sheet and create a simplified 3-column format
    """
    print(f"{Colors.BOLD}📂 Loading workbook...{Colors.END}")
    
    # Load the input workbook
    try:
        source_wb = load_workbook(input_file, data_only=True)
        print(f"{Colors.GREEN}✓{Colors.END} Loaded: {input_file}")
    except Exception as e:
        print(f"{Colors.RED}❌ Error loading file: {str(e)}{Colors.END}")
        return
    
    # Create output workbook
    target_wb = Workbook()
    # Remove default sheet
    if 'Sheet' in target_wb.sheetnames:
        target_wb.remove(target_wb['Sheet'])
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    data_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    print(f"\n{Colors.BOLD}📋 Processing sheets...{Colors.END}")
    print(f"{Colors.CYAN}{'─'*70}{Colors.END}")
    
    total_sheets = len(source_wb.sheetnames)
    total_tracking_numbers = 0
    tracking_summary = []
    
    for sheet_idx, sheet_name in enumerate(source_wb.sheetnames, 1):
        print(f"\n{Colors.BOLD}[{sheet_idx}/{total_sheets}]{Colors.END} Processing: {Colors.YELLOW}{sheet_name}{Colors.END}")
        
        source_sheet = source_wb[sheet_name]
        
        # Find the tracking column
        tracking_col, column_name = find_tracking_column(source_sheet)
        
        if not tracking_col:
            print(f"  {Colors.RED}✗{Colors.END} No tracking column found. Skipping sheet.")
            continue
        
        # Create new sheet in output workbook
        target_sheet = target_wb.create_sheet(title=sheet_name[:31])
        
        # Set headers - use the actual column name found
        headers = [column_name, 'Date Liquidated', 'Logistics Remarks']
        for col_idx, header in enumerate(headers, 1):
            cell = target_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
        
        # Set column widths
        target_sheet.column_dimensions['A'].width = 25  # Tracking column
        target_sheet.column_dimensions['B'].width = 20  # Date Liquidated
        target_sheet.column_dimensions['C'].width = 30  # Logistics Remarks
        
        # Extract tracking numbers from the column
        tracking_numbers = []
        for row_idx in range(2, source_sheet.max_row + 1):
            cell_value = source_sheet.cell(row=row_idx, column=tracking_col).value
            if cell_value and str(cell_value).strip():
                tracking_numbers.append(str(cell_value).strip())
        
        # Write tracking numbers to the new sheet
        if tracking_numbers:
            for row_idx, tracking_num in enumerate(tracking_numbers, 2):
                target_sheet.cell(row=row_idx, column=1).value = tracking_num
                # Apply formatting
                for col_idx in range(1, 4):
                    cell = target_sheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = data_border
                
                # Leave Date Liquidated and Logistics Remarks empty
                target_sheet.cell(row=row_idx, column=2).value = None
                target_sheet.cell(row=row_idx, column=3).value = None
            
            print(f"  {Colors.GREEN}✓{Colors.END} Extracted {Colors.YELLOW}{len(tracking_numbers)}{Colors.END} tracking numbers from '{column_name}'")
            total_tracking_numbers += len(tracking_numbers)
            tracking_summary.append(f"  • {sheet_name}: {len(tracking_numbers)} records (from {column_name})")
        else:
            print(f"  {Colors.YELLOW}⚠{Colors.END} No tracking numbers found in this sheet")
    
    # Save the output file
    print(f"\n{Colors.CYAN}{'─'*70}{Colors.END}")
    print(f"{Colors.BOLD}💾 Saving output file...{Colors.END}")
    
    try:
        target_wb.save(output_file)
        
        # Get file size
        file_size = os.path.getsize(output_file) / (1024 * 1024)
        
        print(f"\n{Colors.GREEN}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}✅ EXTRACTION COMPLETE!{Colors.END}")
        print(f"{Colors.GREEN}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}  📁 Output File  : {Colors.CYAN}{output_file}{Colors.END}")
        print(f"{Colors.BOLD}  📊 File Size    : {Colors.YELLOW}{file_size:.2f} MB{Colors.END}")
        print(f"{Colors.BOLD}  📋 Total Sheets : {Colors.YELLOW}{len(target_wb.sheetnames)}{Colors.END}")
        print(f"{Colors.BOLD}  📈 Total Records: {Colors.YELLOW}{total_tracking_numbers:,}{Colors.END}")
        
        if tracking_summary:
            print(f"\n{Colors.BOLD}  📊 Summary by sheet:{Colors.END}")
            for summary in tracking_summary:
                print(f"    {summary}")
        
        print(f"{Colors.GREEN}{'='*70}{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}✨ All sheets converted to 3-column format!{Colors.END}")
        print(f"{Colors.BOLD}{Colors.GREEN}📝 Column names preserved from source data!{Colors.END}\n")
        
    except Exception as e:
        print(f"{Colors.RED}❌ Error saving file: {str(e)}{Colors.END}")

def main():
    print_banner()
    
    # Get the current working directory
    current_dir = os.getcwd()
    
    # Input file
    input_file = os.path.join(current_dir, 'Shopee Failed Return to Seller (JAN-MARCH).xlsx')
    
    # Output file
    output_file = os.path.join(current_dir, 'Shopee Tracking Numbers (JAN-MARCH).xlsx')
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"{Colors.RED}❌ Error: Input file '{input_file}' not found.{Colors.END}")
        print(f"{Colors.YELLOW}💡 Please run the master file generator first to create the input file.{Colors.END}")
        return
    
    # Record start time
    start_time = datetime.now()
    
    # Run the extraction
    extract_tracking_numbers(input_file, output_file)
    
    # Calculate elapsed time
    elapsed = datetime.now() - start_time
    print(f"{Colors.CYAN}⏱️  Total time: {Colors.YELLOW}{elapsed.total_seconds():.2f}{Colors.END} seconds\n")

if __name__ == "__main__":
    main()